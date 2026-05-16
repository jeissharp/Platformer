import sys
from pygame import *
from openpyxl import load_workbook

# ---------- Константы ----------
win_width = 800
win_height = 500
FPS = 60

# ---------- Глобальные переменные игры ----------
current_level = 0
score = 0
game = True
finish = False

# ---------- Классы ----------
class GameSprite(sprite.Sprite):
    def __init__(self, color, x, y, w, h, spd):
        super().__init__()
        self.image = Surface((w, h))
        self.image.fill(color)
        self.rect = self.image.get_rect()
        self.rect.x = x
        self.rect.y = y
        self.speed = spd

    def reset(self):
        window.blit(self.image, (self.rect.x, self.rect.y))

class Player(GameSprite):
    def __init__(self, color, x, y, w, h, spd):
        super().__init__(color, x, y, w, h, spd)
        self.vel_y = 0
        self.on_ground = False
        self.hp = 3
        self.invincible = 0

    def update(self):
        keys = key.get_pressed()
        if keys[K_LEFT] and self.rect.x > 0:
            self.rect.x -= self.speed
        if keys[K_RIGHT] and self.rect.x < win_width - self.rect.width:
            self.rect.x += self.speed

        if keys[K_SPACE] and self.on_ground:
            self.vel_y = -13
            self.on_ground = False

        self.vel_y += 0.6
        self.rect.y += int(self.vel_y)

        self.on_ground = False
        for p in platforms:
            if self.rect.colliderect(p.rect):
                if self.vel_y > 0:
                    self.rect.bottom = p.rect.top
                    self.vel_y = 0
                    self.on_ground = True
                elif self.vel_y < 0:
                    self.rect.top = p.rect.bottom
                    self.vel_y = 0

        if self.rect.top > win_height:
            self.take_damage()
            self.rect.x = level_data[current_level]['start'][0]
            self.rect.y = level_data[current_level]['start'][1]
            self.vel_y = 0

        if self.invincible > 0:
            self.invincible -= 1
            self.image.set_alpha(100 if self.invincible % 10 < 5 else 255)
        else:
            self.image.set_alpha(255)

    def take_damage(self):
        if self.invincible == 0:
            self.hp -= 1
            self.invincible = 90

class Enemy(GameSprite):
    def __init__(self, color, x, y, w, h, spd, left, right):
        super().__init__(color, x, y, w, h, spd)
        self.left_bound = left
        self.right_bound = right

    def update(self):
        self.rect.x += self.speed
        if self.rect.right >= self.right_bound or self.rect.left <= self.left_bound:
            self.speed *= -1

class Platform(GameSprite):
    def __init__(self, x, y, w, h=40):
        super().__init__((60, 179, 60), x, y, w, h, 0)

class Coin(sprite.Sprite):
    def __init__(self, x, y):
        super().__init__()
        self.image = Surface((20, 20), SRCALPHA)
        draw.circle(self.image, (255, 215, 0), (10, 10), 10)
        self.rect = self.image.get_rect()
        self.rect.x = x
        self.rect.y = y

class Flag(GameSprite):
    def __init__(self, x, y):
        super().__init__((255, 140, 0), x, y, 40, 80, 0)

# ---------- Данные уровней ----------
level_data = [
    {
        'bg': (135, 206, 235),
        'start': (60, 400),
        'platforms': [(0, 460, 800), (150, 360, 150), (350, 280, 150), (550, 200, 150), (100, 180, 120)],
        'coins':     [(230, 330), (425, 250), (625, 170), (160, 150), (400, 430)],
        'enemies':   [(150, 324, 150, 300), (360, 244, 350, 500)],
        'flag':      (730, 400),
    },
    {
        'bg': (135, 206, 235),
        'start': (30, 400),
        'platforms': [(0, 460, 800), (50, 380, 100), (200, 300, 100), (350, 220, 100), (500, 300, 100), (650, 380, 150)],
        'coins':     [(100, 355), (250, 275), (400, 195), (550, 275), (700, 355)],
        'enemies':   [(55, 344, 50, 150), (355, 184, 350, 450), (505, 264, 500, 600)],
        'flag':      (750, 400),
    },
]

CELL = 40
COLOR_MAP = {
    "60B33C": "platform",
    "FFD700": "coin",
    "FF0000": "enemy",
    "FFC000": "flag",
}

def load_level_from_xlsx(path):
    global level_data, current_level
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    platforms = [] 
    coins = []
    enemies = []
    flag_pos = (0, 0)

    for row in ws.iter_rows():
        for cell in row:
            fill = cell.fill
            if not (fill and fill.fill_type == "solid"):
                continue
            color = fill.fgColor.rgb[-6:].upper()
            kind = COLOR_MAP.get(color)

            x = (cell.column - 1) * CELL
            y = (cell.row - 1) * CELL

            if kind == "platform":
                platforms.append((x, y, CELL))  
            elif kind == "coin":
                coins.append((x, y))
            elif kind == "enemy":
                enemies.append((x, y, max(0, x - 3*CELL), x + 3*CELL))
            elif kind == "flag":
                flag_pos = (x, y)

    # Добавляем новый уровень в список
    level_data.append({
        'bg': (135, 206, 235),
        'start': (CELL, win_height - 3*CELL),
        'platforms': platforms,
        'coins': coins,
        'enemies': enemies,
        'flag': flag_pos
    })
    current_level = len(level_data) - 1
    load_level(current_level)

def load_level(index):
    global platforms, coins, enemies, flag
    data = level_data[index]

    platforms = sprite.Group()
    for x, y, w in data['platforms']:
        platforms.add(Platform(x, y, w))

    coins = sprite.Group()
    for x, y in data['coins']:
        coins.add(Coin(x, y))

    enemies = sprite.Group()
    for x, y, l, r in data['enemies']:
        enemies.add(Enemy((220, 0, 0), x, y, 36, 36, 1, l, r))

    flag = sprite.GroupSingle(Flag(*data['flag']))

    player.rect.x = data['start'][0]
    player.rect.y = data['start'][1]
    player.vel_y = 0

# ---------- Инициализация pygame ----------
window = display.set_mode((win_width, win_height))
display.set_caption('Платформер')

font.init()
font1 = font.Font(None, 32)
font2 = font.Font(None, 70)
win_text  = font2.render('ПОБЕДА!',   True, (255, 215, 0))
lose_text = font2.render('GAME OVER', True, (180, 0, 0))

# Создание игровых объектов
player = Player((70, 130, 180), 60, 400, 40, 50, 5)
platforms = sprite.Group()
coins     = sprite.Group()
enemies   = sprite.Group()
flag      = sprite.GroupSingle()

# ---------- Загрузка уровня ----------
if len(sys.argv) > 1:
    load_level_from_xlsx(sys.argv[1])   
else:
    current_level = 0
    load_level(current_level)

clock = time.Clock()

# ---------- Основной игровой цикл (как в shooter.py) ----------
while game:
    for e in event.get():
        if e.type == QUIT:
            game = False

    if not finish:
        window.fill(level_data[current_level]['bg'])

        player.update()
        enemies.update()

        score += len(sprite.spritecollide(player, coins, True)) * 10

        if sprite.spritecollide(player, enemies, False):
            player.take_damage()
            if player.hp <= 0:
                finish = True

        if sprite.spritecollide(player, flag, False):
            current_level += 1
            if current_level >= len(level_data):
                finish = True
            else:
                load_level(current_level)

        platforms.draw(window)
        coins.draw(window)
        enemies.draw(window)
        flag.draw(window)
        player.reset()

        # Жизни (кружочки)
        for i in range(3):
            color = (220, 50, 50) if i < player.hp else (150, 150, 150)
            draw.circle(window, color, (win_width - 30 - i * 35, 20), 12)

        window.blit(font1.render('Счёт: ' + str(score), True, (0, 0, 0)), (10, 10))
        window.blit(font1.render('Уровень: ' + str(current_level + 1), True, (0, 0, 0)), (10, 40))

    else:
        window.fill((30, 30, 30))
        if player.hp <= 0:
            window.blit(lose_text, (win_width // 2 - 180, win_height // 2 - 40))
        else:
            window.blit(win_text,  (win_width // 2 - 140, win_height // 2 - 40))
        window.blit(font1.render('Очки: ' + str(score), True, (200, 200, 200)),
                    (win_width // 2 - 60, win_height // 2 + 40))

    display.update()
    clock.tick(FPS)

quit()